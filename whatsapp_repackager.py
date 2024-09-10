import zipfile
import csv
import json
import re
import os
import shutil
import requests
import pandas as pd
import shortuuid
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference
from pathlib import Path
from slugify import slugify
from datetime import datetime
from collections import defaultdict

""""""""""""""""""""" PARAMETERS TO ADJUST """""""""""""""""""""

EMOJIDESCRIPTION = 'Ask'  # Options: 'Yes' (add descriptions for emoji) / 'No' (don't add descriptions for emoji) / 'Ask' (choose each time you run the script)
API_KEY = 'Ask'           # Options: '[your API-key]' (get it for free at https://emoji-api.com/) / 'Ask' (choose each time you run the script)
LANGUAGE = 'Ask'          # Options: 'en', 'fr', 'nl', 'de', 'es', 'it', 'pt' (language of the application at the time of export) / 'Ask' (choose each time you run the script)
PSYDONYMIZE = 'Ask'       # Options: 'Yes' (to use participants's real names in all output files) / 'No' (to use pseudonymes all output files - the original txt-file will not be modified) / 'Ask' (choose each time you run the script)
FILE_TYPES = 'Ask'        # Options: 'csv', 'xlsx', 'json' / a combination separated by comma's like 'csv, xlsx, json'/ 'Ask' (choose each time you run the script)
OPEN_WHEN_FINISHED = 'Ask'# Options: 'Yes' (open the output folder on completion) / 'No' (don't open the output folder on completion) / 'Ask' (choose each time you run the script)

""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

def extract_zip(zip_path, output_folder):
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(output_folder)

def clean_message_text(text):
    return text.replace("â€Ž", "").strip()

def construct_emoji_dict():
    API_URL = 'https://emoji-api.com/emojis?access_key={}'
    response = requests.get(API_URL.format(api_key))
    if response.status_code == requests.codes.ok:
        data = response.json()
        emoji_dict = {item['character']: item['unicodeName'] for item in data}
        return emoji_dict

def add_emoji_names(emoji_dict, text):
    emoji_pattern = re.compile(
        u'[\U0001F600-\U0001F64F|'
        u'\U0001F300-\U0001F5FF|'
        u'\U0001F680-\U0001F6FF|'
        u'\U0001F700-\U0001F77F|'
        u'\U0001F780-\U0001F7FF|'
        u'\U0001F800-\U0001F8FF|'
        u'\U0001F900-\U0001F9FF|'
        u'\U0001FA00-\U0001FA6F|'
        u'\U0001FA70-\U0001FAFF|'
        u'\U00002700-\U000027BF|'
        u'\U000024C2-\U0001F251]', re.UNICODE)

    def replace(match):
        emoji = match.group(0)
        name = emoji_dict.get(emoji, None)
        if name:
            return f'{emoji} [{name[5:]}]'
        return emoji

    return emoji_pattern.sub(replace, text)

def preprocess_datetime(datetime_str):
    parts = datetime_str.split(' ')
    time_part = parts[1]
    hour, minute = time_part.split(':')
    if len(hour) == 1:
        hour = '0' + hour
    time_part = f"{hour}:{minute}"
    try:
        return f"{parts[0]} {time_part} {parts[2]}"
    except IndexError:
        return f"{parts[0]} {time_part}"

def slugify_filenames_in_folder(folder):
    slugified_filenames = {}
    for file in folder.iterdir():
        if file.is_file():
            slugified_name = slugify(file.stem) + file.suffix
            slugified_path = file.with_name(slugified_name)
            file.rename(slugified_path)
            slugified_filenames[file.name] = slugified_name
    return slugified_filenames

def preprocess_chat_file(txt_file):
    try:
        with open(txt_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
    except FileNotFoundError:
        print("The zip-file appears to have been modified since export. Make sure to use an original zip-file.")
        return

    processed_lines = []
    current_message = ""

    for line in lines:
        match = message_pattern.match(line)
        
        if match:
            if current_message:
                processed_lines.append(current_message.strip())
            
            current_message = line.strip()
        else:
            current_message += " " + line.strip()
    
    if current_message:
        processed_lines.append(current_message.strip())

    with open(txt_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(processed_lines))

def parse_whatsapp_chat(txt_file, attachments_folder, pseudonymize):
    preprocess_chat_file(txt_file)

    with open(txt_file, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    messages = []
    senders = set()

    slugify_filenames_in_folder(attachments_folder)

    message_counts = {}

    if emoji_description:
        emoji_dict = construct_emoji_dict()

    for line in lines:
        match = message_pattern.match(line)
        if match:

            datetime_str, ampm, sender, message = match.groups()
            senders.add(sender)

            # Clean the message text
            message = clean_message_text(message)

            # Add emoji names after emojis
            if emoji_description:
                message = add_emoji_names(emoji_dict, message)
            
            # Convert the datetime string to the desired folder name format: yyyymmddhhmm
            datetime_obj = None

            formatted_str = preprocess_datetime(datetime_str.strip())

            formats = [
                '%d/%m/%Y %H:%M',      # Day/Month/Year Hour:Minute = NL pattern
                '%m/%d/%y, %I:%M %p',  # Month/Day/Year, Hour:Minute AM/PM = EN pattern
            ]

            for fmt in formats:
                try:
                    datetime_obj = datetime.strptime(formatted_str, fmt)
                    break
                except ValueError:
                    continue

            if datetime_obj is None:
                continue

            # Create a standardized folder name
            folder_name = datetime_obj.strftime('%Y%m%d%H%M')

            # Generate a unique ID based on datetime and message count
            message_count = message_counts.get(datetime_str, 0) + 1
            message_counts[datetime_str] = message_count
            message_id = f"{folder_name}_{message_count:02d}"
                      
            # Check if the message references attachments
            if "(" + attachment_indicator + ")" in message:

                # Create a folder named after the message_id
                attachment_folder = attachments_folder / message_id
                if not attachment_folder.exists():
                    attachment_folder.mkdir(parents=True, exist_ok=True)

                # Find all attachment names in the message
                attachment_names_unslugified = re.findall(r"([\S ]+)\s\(bestand bijgevoegd\)", message)
                for attachment_name_unslugified in attachment_names_unslugified:

                    # Create slugified name
                    slugified_name = slugify(Path(attachment_name_unslugified).stem) + Path(attachment_name_unslugified).suffix
                    
                    # Define the source and destination paths
                    source_path = attachments_folder / slugified_name
                    destination_path = attachment_folder / slugified_name
                    
                    if source_path.exists() and source_path.is_file():
                        shutil.move(source_path, destination_path)
                    
                    # Replace references in the message with the slugified filename
                    message = message.replace(f"{attachment_name_unslugified} (bestand bijgevoegd)", f"[{slugified_name}]")
            
                attachment_folder = message_id
            else:
                attachment_folder = ""
                for warning in deleted_message_warnings:
                    if warning.lower() in message.lower():
                        message = "***Deleted message***"
                        break

            # Append the message and attachment folder (if any) to the messages list
            messages.append((message_id, datetime_str, sender, message, attachment_folder))
    
    # Create pseudonym mapping (if applicable)
    pseudonym_mapping = create_pseudonym_mapping(senders)
    
    return messages, sorted(senders), pseudonym_mapping

def create_csv(conversation_name, messages, senders, output_csv, attachments_folder, pseudonym_mapping):
    if pseudonymize:
        # Replace real names with pseudonyms in sender names
        senders = [replace_names_by_pseudonymes(sender, pseudonym_mapping) for sender in senders]

    with open(output_csv, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        
        # Add pseudonymized sender names to header
        header = ['ConversationName', 'MessageID', 'DateTime', 'AttachmentFolder'] + senders
        writer.writerow(header)
        
        for msg in messages:
            message_id, datetime_str, sender, message, attachment_folder = msg
            
            if pseudonymize:
                # Replace real names with pseudonyms in the message
                message = replace_names_by_pseudonymes(message, pseudonym_mapping)
                # Replace real sender names with pseudonyms
                sender = replace_names_by_pseudonymes(sender, pseudonym_mapping)

            row = [conversation_name, message_id, datetime_str, f'=HYPERLINK("{attachments_folder}\\{attachment_folder}")' if attachment_folder else ''] + ['' for _ in senders]

            # Place the message in the correct sender column
            row[4 + senders.index(sender)] = message

            writer.writerow(row)

        if "csv" in file_types:
            if pseudonymize:
                print(f"\u2713 Pseudonymized csv file created: '{output_csv}'")
            else:
                print(f"\u2713 Csv file created: '{output_csv}'")

def create_summary_csv(conversation_name, messages, senders, summary_csv, pseudonym_mapping):
    if not messages:
        return
    
    summary_data = {
        'EarliestMessageDate': [min(msg[1] for msg in messages)],
        'LatestMessageDate': [max(msg[1] for msg in messages)],
        'NumberOfParticipants': [len(senders)],
        'TotalMessages': [len(messages)],
        'TotalAttachments': [sum(1 for msg in messages if msg[4])]
    }

    participant_stats = defaultdict(lambda: {
        'messages': 0,
        'attachments': 0,
        'first_message': None,
        'last_message': None
    })
    for message_id, datetime_str, sender, _, attachment_folder in messages:
        participant_stats[sender]['messages'] += 1
        if attachment_folder:
            participant_stats[sender]['attachments'] += 1
        if participant_stats[sender]['first_message'] is None:
            participant_stats[sender]['first_message'] = datetime_str
        participant_stats[sender]['last_message'] = datetime_str

    for sender, stats in participant_stats.items():
        sender = pseudonym_mapping.get(sender, sender)
        summary_data[f'{sender}_Messages'] = [stats['messages']]
        summary_data[f'{sender}_Attachments'] = [stats['attachments']]
        summary_data[f'{sender}_FirstMessage'] = [stats['first_message']]
        summary_data[f'{sender}_LastMessage'] = [stats['last_message']]
    
    with open(summary_csv, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        for key, value in summary_data.items():
            if pseudonymize:
                writer.writerow([replace_names_by_pseudonymes(key, pseudonym_mapping), replace_names_by_pseudonymes(value[0],pseudonym_mapping)])
            else:
                writer.writerow([key, value[0]])

def create_pseudonym_csv(pseudonym_mapping, output_folder):
    pseudonym_csv = output_folder / "pseudonym_mapping.csv"
    with open(pseudonym_csv, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Real Name', 'Pseudonym'])
        for real_name, pseudonym in pseudonym_mapping.items():
            writer.writerow([real_name, pseudonym])
    print(f"\u2713 Pseudonym mapping csv created: '{pseudonym_csv}'")

def create_pseudonymized_txt(txt_file, pseudonym_mapping, output_folder):
    with open(txt_file, 'r', encoding='utf-8') as file:
        content = file.read()
    pseudonym_txt_file = output_folder / f"{txt_file.stem}_pseudonymized.txt"

    # Replace the names in the entire content using the existing function
    pseudonymized_content = replace_names_by_pseudonymes(content, pseudonym_mapping)

    # Write the pseudonymized content to the output file
    with open(pseudonym_txt_file, 'w', encoding='utf-8') as file:
        file.write(pseudonymized_content)
    
    print(f"\u2713 Pseudonymized text file created: '{pseudonym_txt_file}'")

def create_json_from_csv(csv_file, json_file, pseudonymize=False):
    data = []

    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            conversation_name = row.get("ConversationName")
            message_id = row.get("MessageID")
            date_time = row.get("DateTime")
            attachment_folder = row.get("AttachmentFolder")

            # Iterate over each participant's message field
            for participant, message in row.items():
                if participant not in ["ConversationName", "MessageID", "DateTime", "AttachmentFolder"] and message.strip():
                    data.append({
                        "ConversationName": conversation_name,
                        "MessageID": message_id,
                        "DateTime": date_time,
                        "Name": participant,
                        "Message": message
                    })

    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

    if pseudonymize:
        print(f"\u2713 Pseudonymised json file created: '{json_file}'")
    else:
        print(f"\u2713 Json file created: '{json_file}'")

def create_excel_from_csv(csv_file, excel_file, summary_csv):
    
    # Load CSV file
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)

    # Extract the headers and sender names
    headers = rows[0]
    senders = headers[4:]
    sender_colors = assign_colors_to_senders(senders)
    
    # Create a new Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WhatsApp Chat"

    # Write headers to the Excel file
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Define font style for deleted messages
    deleted_message_font = Font(color="808080", italic=True)

    # Write rows to the Excel file
    for row_num, row in enumerate(rows[1:], 2):  # Skip the header
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            
            # Apply color formatting based on sender
            if col_num > 4:  # Message columns start from the 5th column
                sender = headers[col_num-1]
                if cell_value:

                    # Apply color fill for messages
                    cell.fill = sender_colors.get(sender, PatternFill())

                    # Special formatting for deleted messages
                    if "***Deleted message***" in cell_value:
                        cell.font = deleted_message_font
                        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Apply color to the attachment folder cell if it exists
            if col_num == 4 and cell_value:
                sender_index = next((i for i, value in enumerate(row[4:], 4) if value), None)
                if sender_index is not None:
                    attachment_color = sender_colors.get(headers[sender_index])
                    cell.fill = attachment_color

    # Add a new worksheet for summary statistics
    ws_summary = wb.create_sheet(title="Summary")
    df_summary = pd.read_csv(summary_csv)
    for r in dataframe_to_rows(df_summary, index=False, header=True):
        ws_summary.append(r)

    # Identify rows in column A that contain "_Messages"
    message_rows = []
    for row in ws_summary.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if '_Messages' in cell.value:
            message_rows.append(cell.row)

    # Convert the message count cells to integers (if they are formatted as text)
    for row in message_rows:
        message_count_cell = ws_summary.cell(row=row, column=2)
        try:
            message_count_cell.value = int(message_count_cell.value)
        except ValueError:
            pass  # In case the value isn't a valid integer, do nothing

    # Create a helper table for pie chart data
    temp_table_start_row = 18
    temp_table_start_col = 6
    ws_summary.cell(row=temp_table_start_row - 1, column=temp_table_start_col, value="Participant")
    ws_summary.cell(row=temp_table_start_row - 1, column=temp_table_start_col + 1, value="Number of Messages")
    
    for index, row in enumerate(message_rows):
        participant_name = ws_summary.cell(row=row, column=1).value
        message_count = ws_summary.cell(row=row, column=2).value
        ws_summary.cell(row=temp_table_start_row + index, column=temp_table_start_col, value=participant_name)
        ws_summary.cell(row=temp_table_start_row + index, column=temp_table_start_col + 1, value=message_count)

    # Create a pie chart for message distribution by participant
    pie_chart = PieChart()
    
    # Define the data range for the pie chart
    data = Reference(ws_summary, min_col=temp_table_start_col + 1, min_row=temp_table_start_row, max_col=temp_table_start_col + 1, max_row=temp_table_start_row + len(message_rows) - 1)
    labels = Reference(ws_summary, min_col=temp_table_start_col, min_row=temp_table_start_row, max_row=temp_table_start_row + len(message_rows) - 1)
    
    pie_chart.add_data(data, titles_from_data=False)
    pie_chart.set_categories(labels)
    pie_chart.title = "Message Distribution by Participant"
    
    # Add the pie chart to the "Summary" worksheet
    ws_summary.add_chart(pie_chart, "E2")

    # Save the Excel file
    wb.save(excel_file)
    if pseudonymize:
        print(f"\u2713 Pseudonymized Excel file created: '{excel_file}'")
    else:
        print(f"\u2713 Excel file created: '{excel_file}'")

def assign_colors_to_senders(senders):
    color_palette = [
        "FFCCCC", "CCFFCC", "CCCCFF", "FFFFCC", "FFCCFF", "CCFFFF", "FFD700", 
        "FF69B4", "87CEFA", "98FB98", "FFDAB9", "FFA07A", "D3D3D3"
    ]
    sender_colors = {}
    for i, sender in enumerate(senders):
        color = color_palette[i % len(color_palette)]
        sender_colors[sender] = PatternFill(start_color=color, end_color=color, fill_type="solid")
    return sender_colors

def process_whatsapp_zip(zip_path, pseudonymize):
    zip_path = Path(zip_path)
    output_folder = zip_path.parent / zip_path.stem
    
    # Check if the output folder already exists
    if output_folder.exists():
        user_choice = input(f"\u2022 The output folder already exists. Do you want to delete '{output_folder}' and continue? (yes/no): ").strip().lower()
        if user_choice == 'yes':
            shutil.rmtree(output_folder)
            print(f"\u2713 Deleted folder '{output_folder}'")
        else:
            print("Operation canceled by the user.")
            return
    
    # Extract conversation name from the ZIP file name (without extension)
    conversation_name = zip_path.stem
    
    # Create the output folder
    output_folder.mkdir(exist_ok=True)
    print(f"\u2713 Created folder '{output_folder}'")
    
    # Extract the ZIP file
    extract_zip(zip_path, output_folder)
    
    # Find the txt file and attachments
    txt_file = output_folder / f"{zip_path.stem}.txt"
    attachments_folder = output_folder / 'attachments'
    attachments_folder.mkdir(exist_ok=True)

    # Move all files except the txt file to the attachments folder
    folder = output_folder.resolve()
    for item in folder.iterdir():
        if item.name != f"{zip_path.stem}.txt" and item.is_file():
            item.rename(attachments_folder / item.name)
    print(f"\u2713 Attachments moved to dedicated folder: '{output_folder}'")

    # Slugify all filenames in the attachments folder and create a map for reference
    slugify_filenames_in_folder(attachments_folder)

    # Parse the WhatsApp chat
    messages, senders, pseudonym_mapping = parse_whatsapp_chat(txt_file, attachments_folder, pseudonymize)

    suffix = "_pseudonymized" if pseudonymize else ""

    # Create the CSV file
    output_csv = output_folder / f"{zip_path.stem}{suffix}.csv"
    create_csv(conversation_name, messages, senders, output_csv, attachments_folder, pseudonym_mapping)
    
    # Create the summary CSV file
    output_summary_csv = output_folder / f"{zip_path.stem}_summary{suffix}.csv"
    create_summary_csv(conversation_name, messages, senders, output_summary_csv, pseudonym_mapping)
    
    # Create the JSON file
    if "json" in file_types:
        output_json = output_folder / f"{zip_path.stem}{suffix}.json"
        create_json_from_csv(output_csv, output_json)

    # Create the Excel file with pie chart
    if "xlsx" in file_types:
        excel_file = output_folder / f"{zip_path.stem}{suffix}.xlsx"
        create_excel_from_csv(output_csv, excel_file, output_summary_csv)

    if not "csv" in file_types:
        os.remove(output_csv)

    if pseudonymize:
        create_pseudonymized_txt(txt_file, pseudonym_mapping, output_folder)
        create_pseudonym_csv(pseudonym_mapping, output_folder)
    
    print(f"\033[92mProcessing complete. Output saved to '{output_folder}'\033[0m")

    if pseudonymize:
        print("\033[93mWarning: Although the sender's names have been pseudonymized, other names are not and senders may still be identifiable based on metadata and/or message content.\033[0m")

    open_output_folder = False
    if OPEN_WHEN_FINISHED == "yes":
        open_output_folder = True
    elif OPEN_WHEN_FINISHED != "no":
        open_output_folder_valid_input = False
        while not open_output_folder_valid_input:
            open_output_folder_input = input("Open ouput folder? (yes/no): ").strip().lower()
            if open_output_folder_input == "yes":
                open_output_folder = True
                open_output_folder_valid_input = True
            elif open_output_folder_input == "no":
                open_output_folder_valid_input = True
            else:
                print("Invalid input.")
    if open_output_folder:
        os.startfile(output_folder)

def create_pseudonym_mapping(senders):
    pseudonym_mapping = {}
    if pseudonymize:
        for sender in senders:
            if sender not in pseudonym_mapping:
                pseudonym_mapping[sender] = shortuuid.ShortUUID().random(length=6)
    return pseudonym_mapping            
            
def replace_names_by_pseudonymes(text, mapping):
    for real_name, pseudonym in mapping.items():
        text = str(text).replace(real_name, pseudonym)
    return text

if __name__ == "__main__":
    message_pattern = re.compile(r"(\d{1,2}/\d{1,2}/\d{2,4},?\s*\d{1,2}:\d{2}\s*([ap]m\s)?)- (.+?): (.+)", re.IGNORECASE)
    
    zip_file_path = input("\u2022 Enter the path to the WhatsApp ZIP file: ").strip().replace('"','')
    
    allowed_languages = ['EN', 'FR', 'NL', 'DE', 'ES', 'IT', 'PT']
    if LANGUAGE.strip().upper() in allowed_languages:
        language = LANGUAGE.strip().upper()
    else:
        valid_language_input = False
        while not(valid_language_input):
            language_input = input("\u2022 No language was set in the script. What was the interface language of the export? (en/fr/nl/de/es/it/pt): ").strip().upper()
            if language_input in allowed_languages:
                language = language_input
                valid_language_input = True
            else:
                print("Invalid input.")
    
    attachment_indicator = ""
    deleted_message_warnings = []
    if language.strip().upper() == 'EN':  # Translation confirmed 28.8.2024
        attachment_indicator = "file attached"
        deleted_message_warnings = ["This message was deleted", "You deleted this message"]
    elif language.strip().upper() == 'FR':  # Translation confirmed 28.8.2024
        attachment_indicator = "fichier joint"
        deleted_message_warnings = ["Ce message a été supprimé", "Vous avez supprimé ce message"]
    elif language.strip().upper() == 'NL':  # Translation confirmed 28.8.2024
        attachment_indicator = "bestand bijgevoegd"
        deleted_message_warnings = ["Dit bericht is verwijderd", "U hebt dit bericht verwijderd"]
    elif language.strip().upper() == 'DE':  # Translation not confirmed
        attachment_indicator = "Dateianhang"
        deleted_message_warnings = ["Diese Nachricht wurde gelöscht", "Sie haben diese Nachricht gelöscht"]
    elif language.strip().upper() == 'ES':  # Translation not confirmed
        attachment_indicator = "archivo adjunto"
        deleted_message_warnings = ["Este mensaje fue eliminado", "Has eliminado este mensaje"]
    elif language.strip().upper() == 'IT':  # Translation not confirmed
        attachment_indicator = "file allegato"
        deleted_message_warnings = ["Questo messaggio è stato eliminato", "Hai eliminato questo messaggio"]
    elif language.strip().upper() == 'PT':  # Translation not confirmed
        attachment_indicator = "arquivo anexado"
        deleted_message_warnings = ["Esta mensagem foi apagada", "Você apagou esta mensagem"]
    else:
        raise ValueError(f"Unsupported language code: {language}")

    file_type_valid_input = False
    allowed_file_types = ['csv', 'xlsx', 'json']
    if FILE_TYPES.strip().lower() != "ask":
        file_types = FILE_TYPES.split(",")
        file_types = [type.strip().lower() for type in file_types]
        invalid_type = False
        for type in file_types:
            if type not in allowed_file_types:
                invalid_type = True
                print("Invalid file types parameter. Check if all types in FILE_TYPES are allowed.")
        if not invalid_type:
            file_type_valid_input = True
    else:
        while not file_type_valid_input:    
            file_types_input = input("\u2022 Which files do you want to generate? (csv, xlsx, json, multiple separated by commas, or enter for all): ").strip().lower()
            if not file_types_input:
                file_types = ['csv', 'xlsx', 'json']
                file_type_valid_input = True
            else:
                file_types_input = file_types_input.split(",")
                file_types_input = [type.strip().lower() for type in file_types_input]
                invalid_type = False
                for type in file_types_input:
                    if type not in allowed_file_types:
                        invalid_type = True
                        print("Invalid input.")
                if not invalid_type:
                    file_types = file_types_input
                    file_type_valid_input = True

    if file_type_valid_input:

        if PSYDONYMIZE.strip().lower() == "yes":
            pseudonymize = True
        elif PSYDONYMIZE.strip().lower() == "no":
            pseudonymize = False
        else:
            pseudonymize_valid_input = False
            while not pseudonymize_valid_input:
                pseudonymize_input = input("\u2022 Would you like to pseudonymize senders' names? (yes/no): ").lower()
                if pseudonymize_input == 'yes':
                    pseudonymize = True
                    pseudonymize_valid_input = True
                elif pseudonymize_input == "no":
                    pseudonymize = False
                    pseudonymize_valid_input = True
                else:
                    print("Invalid input")

        if EMOJIDESCRIPTION.strip().lower() == "no":
            emoji_description = False
        elif EMOJIDESCRIPTION.strip().lower() == "yes":
            emoji_description = True
        else:
            emoji_description_valid_input = False
            while not emoji_description_valid_input:
                emoji_description_input = input("\u2022 Would you like to add emoji descriptions to the output files (requires api key)? (yes/no): ").lower()
                if emoji_description_input == 'yes':
                    emoji_description = True
                    emoji_description_valid_input = True
                elif emoji_description_input == "no":
                    emoji_description = False
                    emoji_description_valid_input = True
                else:
                    print("Invalid input.")

        api_key = ""
        if emoji_description:
            api_pattern = r"^[0-9a-fA-F]{40}$"
            if re.match(api_pattern, API_KEY) is None:
                proceed_with_no_key_valid_input = False
                while not proceed_with_no_key_valid_input:
                    proceed_with_no_key = input("\u2022 No valid emoji api key was found in the script. Proceed without adding emoji descriptions? (yes/no): ").strip().lower()
                    if proceed_with_no_key == 'yes':
                        emoji_description = False
                        proceed_with_no_key_valid_input = True
                    elif proceed_with_no_key == 'no':
                        enter_key_valid_input = False
                        while not enter_key_valid_input:
                            api_key_input = input("\u2022 Enter a valid key or hit enter to abort the operation: ").strip().lower()
                            if not api_key_input:
                                print("Operation canceled by the user.")
                                enter_key_valid_input = True
                                proceed_with_no_key_valid_input = True
                            elif re.match(api_pattern, api_key_input) is None:
                                print("Invalid key.")
                            else:
                                api_key = api_key_input
                                enter_key_valid_input = True
                                proceed_with_no_key_valid_input = True
            else:
                api_key = API_KEY              

        if api_key or not(emoji_description):
            process_whatsapp_zip(zip_file_path, pseudonymize)



""" AI assistance was used during the development of this script. """
